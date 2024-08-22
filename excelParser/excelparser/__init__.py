from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import re

# SOURCE  SHEET 
SHEET_COVER = "COVER"
SHEET_CONSIGNMENT_DATA = "Consignment Data"

# DESTINATION SHEET 
SHEET_STANDARD_FREIGHT_IMPORT_TEMPLATE = "Standard Freight Import Templat"

# SOURCE COLUMN 
CONSIGNMENT_REFERENCE = "Consignment Reference"
CONSIGNMENT_ID = "Consignment ID"

# DESTINATION COLUMN 
ORDER_NO = "Order No"
BOOKING_NO = "Booking No"


def ProcessConsignmentRefValue(value):
    return re.split(r'[/-]',value)[0]


def OpenBrowseDialog(prompt):
    Tk().withdraw()  
    file_path = askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    return file_path


def IterateColumn(columnName,sheet,ref):
    colDataList = []
    id_col = None
    for cell in sheet[1]:
        if cell.value == columnName:
            id_col = cell.column
            break

    if id_col is None:
        print(f"No {columnName} column found.")
    else:
        for row in sheet.iter_rows(min_row=2, min_col=id_col, max_col=id_col, values_only=True):
            id_value = row[0]
            if id_value:
                if ref:
                    id_value = ProcessConsignmentRefValue(id_value)
                colDataList.append(id_value)
    return colDataList


def LoadExcelFile():
    excelFile = OpenBrowseDialog("Select the source Excel file")
    return excelFile

def ProcessConsignmentId(ConsignmentIdColumn,sheetName):
    sheet_name = sheetName
    workbook = load_workbook(filename=sourceExcel)
    sheet = workbook[sheet_name]
    return IterateColumn(ConsignmentIdColumn,sheet)

def ProcessConsignmentReference(ConsignmentRefColumn,sheetName):
    sheet_name = sheetName
    workbook = load_workbook(filename=sourceExcel)
    sheet = workbook[sheet_name]
    IterateColumn(ConsignmentRefColumn,sheet)

def ProcessColumn(columnName,sheetName,ref):
    print(f"Processing column: {columnName}, Sheet: {sheetName}")
    sheet_name = sheetName
    workbook = load_workbook(filename=sourceExcel)
    sheet = workbook[sheet_name]
    return IterateColumn(columnName,sheet,ref)

def saveToFile(data, columnName, sheetName):
    ws = destinationWB[sheetName]
    col = None 
    for cell in ws[1]:
        if cell.value == columnName:
            col = cell.column
            break
    ws.cell(row=1, column=col, value=columnName)
    for indx, value in enumerate(data, start=2):
        ws.cell(row=indx, column=col, value=value)

    destinationWB.save(destinationExcel)

def StartProcess(sourceColumnName,sourceSheetName,destinationColumnName,destinationSheetName,regExFlag):
    consignmentIdData = ProcessColumn(sourceColumnName,sourceSheetName,regExFlag)
    saveToFile(consignmentIdData,destinationColumnName,destinationSheetName)


if __name__ == "__main__":
    sourceExcel = LoadExcelFile()
    destinationExcel = LoadExcelFile()
    destinationWB = load_workbook(filename=destinationExcel)

    StartProcess(CONSIGNMENT_ID,SHEET_CONSIGNMENT_DATA,BOOKING_NO,SHEET_STANDARD_FREIGHT_IMPORT_TEMPLATE,False)
    StartProcess(CONSIGNMENT_REFERENCE,SHEET_CONSIGNMENT_DATA,ORDER_NO,SHEET_STANDARD_FREIGHT_IMPORT_TEMPLATE,True)