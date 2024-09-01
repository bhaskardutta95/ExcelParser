from openpyxl import load_workbook,utils
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import re


def OpenBrowseDialog(prompt):
    Tk().withdraw()  
    file_path = askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    return file_path

def SaveFile(wb, file_path):
    wb.save(file_path)


def readOrderIdsFromExcel(sheet, orderIdColumnName):
    orderIds = []

    # Find the column index for the "OrderId" column
    orderIdColumnIndex = None
    for cell in sheet[1]:  # Assuming the first row contains headers
        if cell.value == orderIdColumnName:
            orderIdColumnIndex = cell.column
            break

    if orderIdColumnIndex is None:
        # raise ValueError(f"Column '{orderIdColumnName}' not found in the sheet.")
        print(f"Column '{orderIdColumnName}' not found in the sheet.")

    # Iterate through the rows and collect the OrderIds
    for row in sheet.iter_rows(min_row=2):  # Skip the header row
        orderId = row[orderIdColumnIndex - 1].value  # -1 because row is 0-indexed but columns are 1-indexed
        if orderId is not None:
            orderIds.append(orderId)

    return orderIds


def find_column_index(sheet, column_name):
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value == column_name:
            return col
    return None

def FindFirstEmptyRow(sheet):
    for row in range(1, sheet.max_row + 1):
        if not any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row
    return sheet.max_row + 1  # If no empty row is found, append at the end

def CopyColumns(destination_wb, column_mapping, source_sheet, destination_sheet):
    # Find the columns in the Source and Destination sheets based on the provided mapping
    source_columns = {src_col: find_column_index(source_sheet, src_col) for src_col in column_mapping.keys()}
    missing_source_columns = [col for col, idx in source_columns.items() if idx is None]

    destination_columns = {dest_col: find_column_index(destination_sheet, dest_col) for dest_col in column_mapping.values()}
    missing_destination_columns = [col for col, idx in destination_columns.items() if idx is None]

     # Report missing columns
    if missing_source_columns or missing_destination_columns:
        if missing_source_columns:
            print(f"Source columns not found: {', '.join(missing_source_columns)}")
        if missing_destination_columns:
            print(f"Destination columns not found: {', '.join(missing_destination_columns)}")
        return

    # Find the first empty row in the Destination sheet
    start_row = FindFirstEmptyRow(destination_sheet)

    # Iterate through the rows in the Source sheet and copy data
    for row in range(2, source_sheet.max_row + 1):  # Start from 2 if there's a header row
        data_to_copy = {dest_col: source_sheet.cell(row=row, column=source_columns[src_col]).value
                        for src_col, dest_col in column_mapping.items()}

        # Write the data to the Destination sheet
        for dest_col, value in data_to_copy.items():
            destination_sheet.cell(row=start_row, column=destination_columns[dest_col], value=value)
        start_row += 1  # Move to the next row

    # Save the updated Destination workbook
    print("Data copied successfully!")
    return destination_wb

def SearchRefAndBreakdown(sheet,referenceNumberList):
    breakdownColumn = 'L'
    refAndBreakdownList = {}
    referencesNotFound = set(referenceNumberList)  # Initially assume all words are not found

    # Iterate over all cells in the sheet
    for row in sheet.iter_rows(min_row=2):  # Assuming the first row contains headers
        for cell in row:
            if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string
                for refNumber in referenceNumberList:
                    if refNumber in cell.value:
                        price = sheet.cell(row=cell.row, column=utils.column_index_from_string(breakdownColumn)).value
                        refAndBreakdownList[refNumber] = price
                        referencesNotFound.discard(refNumber)  # Remove the refNumber from referencesNotFound if found

    return refAndBreakdownList, list(referencesNotFound)

def updateGrossAmounts(workbook,sheet,referenceGrossMap, filePath):
    orderNoColumnIndex = None
    grossColumnIndex = None
    for cell in sheet[1]:  # Assuming the first row contains headers
        if cell.value and isinstance(cell.value, str):
            if "order" in cell.value.lower() and "no" in cell.value.lower():
                orderNoColumnIndex = cell.column
            if "gross" in cell.value.lower():
                grossColumnIndex = cell.column

    if orderNoColumnIndex is None or grossColumnIndex is None:
        raise ValueError("Required columns 'Order No' or 'Gross' not found in the sheet.")

    # Iterate through the rows and update the Gross column based on matching Order No
    for row in sheet.iter_rows(min_row=2):  # Assuming the first row contains headers
        orderNo = row[orderNoColumnIndex - 1].value  # Adjust for 0-based indexing
        if orderNo in referenceGrossMap:
            grossText = referenceGrossMap[orderNo]
            grossAmount = extractDollarValue(grossText)
            if grossAmount:
                sheet.cell(row=row[0].row, column=grossColumnIndex).value = float(grossAmount)

    workbook.save(filePath)
    print("Gross amounts successfully updated in the Excel sheet.")

def extractDollarValue(text):
    match = re.search(r'=\s*\$([\d,]+\.\d{2})', text)
    if match:
        return match.group(1).replace(',', '')
    return None


if __name__ == "__main__":
    sourceExcel = OpenBrowseDialog("Select the source Excel file")
    destinationExcel = OpenBrowseDialog("Select the destination Excel file")

    sourceSheet = "Consignment Data"
    destinationSheet = "Standard Freight Import Templat"
    ConsignmentsAndManifests = "Consignments and Manifests"
    OrderNo = "Order No"

    source_wb = load_workbook(sourceExcel)
    source_sheet = source_wb[sourceSheet]

    destination_wb = load_workbook(destinationExcel)
    destination_sheet = destination_wb[destinationSheet]

    ConsignmentsAndManifestsSheet = source_wb[ConsignmentsAndManifests]

    column_mapping = {
        'Consignment ID': 'Booking No',
        'Consignment Reference': 'Order No',
        'Total SPC': 'Charge Qty'
    }

    destinationWB = CopyColumns(destination_wb, column_mapping, source_sheet, destination_sheet)
    SaveFile(destinationWB, destinationExcel)

    orderNumbers = readOrderIdsFromExcel(destination_sheet, OrderNo)
 
    refAndBreakdownList = SearchRefAndBreakdown(ConsignmentsAndManifestsSheet,orderNumbers)

    updateGrossAmounts(destination_wb,destination_sheet,refAndBreakdownList[0], destinationExcel)

